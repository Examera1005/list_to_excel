#!/usr/bin/env python3
"""
Générateur de Fichiers Élèves - Interface Terminal (TUI)
Compatible macOS / Linux / Windows

Fonctionnalités avancées :
- Contrôle strict des saisies utilisateur (boucle d'erreur en cas d'entrée invalide)
- Gestionnaire de sélection MULTIPLE de dossiers de destination avec navigateur
"""

import os
import sys
import glob
import subprocess

def ensure_openpyxl():
    """Vérifie si openpyxl est installé et propose de l'installer si nécessaire."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("===============================================================")
        print(" ⚠️ DÉPENDANCE MANQUANTE : openpyxl")
        print("===============================================================")
        print("La bibliothèque 'openpyxl' est nécessaire pour lire et modifier")
        print("les fichiers Excel (.xlsx).\n")
        
        choix = input("👉 Souhaitez-vous l'installer automatiquement maintenant ? [O/n] : ").strip().lower()
        if choix in ['', 'o', 'oui', 'y', 'yes']:
            print("\n⏳ Installation de 'openpyxl' via pip...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            except Exception:
                try:
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"])
                except Exception as err:
                    print(f"\n❌ Échec de l'installation automatique : {err}")
                    print("💡 Veuillez installer la bibliothèque manuellement avec :")
                    print("   pip3 install openpyxl")
                    sys.exit(1)
                    
            print("✅ 'openpyxl' a été installé avec succès !\n")
            import openpyxl
            return openpyxl
        else:
            print("\n❌ Impossible de continuer sans 'openpyxl'.")
            sys.exit(1)

openpyxl = ensure_openpyxl()

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

def get_valid_input(prompt_text, valid_options, default=None):
    """
    Demande une saisie à l'utilisateur et boucle tant que la réponse n'est pas valide.
    valid_options: liste de chaînes (ex: ['1', '2', 'M', 'S'])
    """
    valid_upper = [str(o).upper() for o in valid_options]
    while True:
        saisie = input(prompt_text).strip()
        if not saisie and default is not None:
            return str(default).upper()
        if saisie.upper() in valid_upper:
            return saisie.upper()
        print(f"❌ Saisie invalide ('{saisie}'). Options autorisées : {', '.join(map(str, valid_options))}" + 
              (f" [Appuyez sur Entrée pour '{default}']" if default is not None else ""))

def get_native_file_picker(title="Choisir un fichier", file_types=""):
    """Tente d'ouvrir la fenêtre native macOS pour choisir un fichier."""
    if sys.platform == "darwin":
        try:
            cmd = f'osascript -e "POSIX path of (choose file with prompt \\"{title}\\")"'
            res = subprocess.check_output(cmd, shell=True).decode('utf-8').strip()
            if res:
                return res
        except Exception:
            pass
    return None

def get_native_folder_picker(title="Choisir un dossier"):
    """Tente d'ouvrir la fenêtre native macOS pour choisir un dossier."""
    if sys.platform == "darwin":
        try:
            cmd = f'osascript -e "POSIX path of (choose folder with prompt \\"{title}\\")"'
            res = subprocess.check_output(cmd, shell=True).decode('utf-8').strip()
            if res:
                return res
        except Exception:
            pass
    return None

def parse_student_list(liste_path):
    """Extrait les élèves depuis un fichier Excel (.xlsx) ou Texte (.txt)."""
    eleves = []
    
    if liste_path.endswith(".xlsx"):
        wb = openpyxl.load_workbook(liste_path, data_only=True)
        ws = wb.active
        
        header_row = None
        for r in range(1, 30):
            vals = [str(ws.cell(row=r, column=c).value or '').strip() for c in range(1, 15)]
            if 'Nom' in vals and ('Prénom' in vals or 'Prenom' in vals):
                header_row = r
                break
                
        if not header_row:
            print("❌ En-têtes 'Nom' et 'Prénom' introuvables dans le fichier Excel.")
            return []
            
        nom_c = next(c for c in range(1, 15) if str(ws.cell(row=header_row, column=c).value).strip().lower() == 'nom')
        prenom_c = next(c for c in range(1, 15) if str(ws.cell(row=header_row, column=c).value).strip().lower() in ['prénom', 'prenom'])
        group_c = next((c for c in range(1, 15) if str(ws.cell(row=header_row, column=c).value).strip().lower() in ['groupe', 'classe']), None)
        
        for r in range(header_row + 1, ws.max_row + 1):
            nom = ws.cell(row=r, column=nom_c).value
            prenom = ws.cell(row=r, column=prenom_c).value
            grp = ws.cell(row=r, column=group_c).value if group_c else ''
            
            if nom and prenom:
                eleves.append({
                    "nom": str(nom).strip(),
                    "prenom": str(prenom).strip(),
                    "classe": str(grp).strip() if grp else ""
                })
    else:
        with open(liste_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split()
                    if len(parts) >= 2:
                        prenom = parts[0]
                        nom = " ".join(parts[1:])
                    else:
                        prenom = line
                        nom = ""
                    eleves.append({"nom": nom, "prenom": prenom, "classe": ""})
                    
    return eleves

def multi_folder_browser():
    """
    Navigateur interactif permettant de parcourir les sous-dossiers,
    d'en créer et de cocher/décocher plusieurs dossiers de destination.
    """
    current_dir = os.getcwd()
    selected_folders = set()
    
    # Par défaut, ajouter './fichiers_eleves'
    default_target = os.path.abspath("fichiers_eleves")
    selected_folders.add(default_target)

    while True:
        clear_screen()
        print("===============================================================")
        print(" 📁 NAVIGATEUR ET SÉLECTION MULTIPLE DE DOSSIERS CIBLES")
        print("===============================================================")
        print(f"📍 Dossier actuel de navigation : {current_dir}\n")
        
        print("🎯 Dossier(s) actuellement sélectionné(s) pour la génération :")
        if not selected_folders:
            print("   (Aucun dossier sélectionné !)")
        else:
            for sf in sorted(selected_folders):
                print(f"   [✓] {sf}")
        print("---------------------------------------------------------------")
        
        # Lister les sous-dossiers dans current_dir
        try:
            entries = sorted([d for d in os.listdir(current_dir) if os.path.isdir(os.path.join(current_dir, d)) and not d.startswith('.')])
        except PermissionError:
            entries = []
            print("⚠️ Accès refusé à ce dossier.")

        print("\nSous-dossiers disponibles :")
        print("  [..] ⬆️  Remonter d'un dossier parent")
        print("  [+]  ➕ Créer un nouveau sous-dossier ici")
        if sys.platform == "darwin":
            print("  [M]  🖥️  Ajouter un dossier via le sélecteur macOS")
        print("  [OK] 🚀 VALIDER LA SÉLECTION et continuer")
        print("")
        
        dir_map = {}
        for idx, folder_name in enumerate(entries, 1):
            full_path = os.path.abspath(os.path.join(current_dir, folder_name))
            isChecked = "✓" if full_path in selected_folders else " "
            print(f"  [{idx}] [{isChecked}] 📂 {folder_name}")
            dir_map[str(idx)] = (folder_name, full_path)
            
        print("\nCommandes possibles :")
        print("  • Tapez un N° pour entrer dans un sous-dossier (ex: 1)")
        print("  • Tapez C<N°> pour Cocher/Décocher un dossier (ex: C1 pour cocher/décocher le dossier 1)")
        print("  • Tapez 'OK' pour terminer la sélection.")
        
        cmd = input("\n👉 Votre choix : ").strip().upper()
        
        if cmd == 'OK':
            if not selected_folders:
                print("❌ Veuillez sélectionner au moins un dossier de destination !")
                input("Appuyez sur Entrée pour continuer...")
                continue
            return list(selected_folders)
        elif cmd == '..':
            parent = os.path.dirname(current_dir)
            if parent and parent != current_dir:
                current_dir = parent
        elif cmd == '+':
            new_name = input("👉 Nom du nouveau sous-dossier : ").strip()
            if new_name:
                new_path = os.path.join(current_dir, new_name)
                os.makedirs(new_path, exist_ok=True)
                selected_folders.add(os.path.abspath(new_path))
                print(f"✅ Dossier créé et sélectionné : {new_path}")
        elif cmd == 'M' and sys.platform == "darwin":
            mac_folder = get_native_folder_picker("Choisir un dossier de destination")
            if mac_folder:
                selected_folders.add(os.path.abspath(mac_folder))
        elif cmd.startswith('C') and cmd[1:].isdigit():
            idx_str = cmd[1:]
            if idx_str in dir_map:
                _, full_p = dir_map[idx_str]
                if full_p in selected_folders:
                    selected_folders.remove(full_p)
                else:
                    selected_folders.add(full_p)
            else:
                print(f"❌ Numéro de dossier invalide : {idx_str}")
                input("Appuyez sur Entrée...")
        elif cmd in dir_map:
            # Naviguer dans le sous-dossier
            _, full_p = dir_map[cmd]
            current_dir = full_p
        else:
            print(f"❌ Commande non reconnue : '{cmd}'")
            input("Appuyez sur Entrée pour réessayer...")

def main():
    clear_screen()
    print("===============================================================")
    print("  🎓 GÉNÉRATEUR DE DOCUMENTS ÉLÈVES (Template Excel)")
    print("===============================================================\n")

    # -------------------------------------------------------------------------
    # 1. SÉLECTION DU FICHIER TEMPLATE (.xlsx)
    # -------------------------------------------------------------------------
    print("📌 1. FICHIER TEMPLATE EXCEL")
    templates_locaux = glob.glob("*.xlsx")
    template_path = ""

    if templates_locaux:
        print("Fichiers Excel trouvés dans le dossier actuel :")
        valid_opts = []
        def_idx = 1
        for idx, f in enumerate(templates_locaux, 1):
            print(f"  [{idx}] {f}")
            valid_opts.append(str(idx))
            if "copie" in f.lower() or "template" in f.lower():
                def_idx = idx

        if sys.platform == "darwin":
            print("  [M] Ouvrir le sélecteur macOS")
            valid_opts.append("M")
        print("  [S] Saisir le chemin manuellement")
        valid_opts.append("S")
        
        choix = get_valid_input(f"\n👉 Choix (défaut={def_idx}) : ", valid_opts, default=def_idx)
        
        if choix == 'M':
            template_path = get_native_file_picker("Sélectionnez votre fichier template Excel")
        elif choix == 'S':
            while True:
                template_path = input("Chemin du fichier template : ").strip()
                if template_path and os.path.exists(template_path):
                    break
                print(f"❌ Fichier introuvable : '{template_path}'. Réessayez.")
        else:
            template_path = templates_locaux[int(choix) - 1]
    else:
        print("Aucun fichier .xlsx dans le dossier courant.")
        if sys.platform == "darwin":
            template_path = get_native_file_picker("Sélectionnez votre fichier template Excel")
        while not template_path or not os.path.exists(template_path):
            template_path = input("👉 Entrez le chemin du fichier template.xlsx : ").strip()
            if not os.path.exists(template_path):
                print(f"❌ Fichier introuvable : '{template_path}'")

    print(f"✅ Template sélectionné : {template_path}\n")

    # -------------------------------------------------------------------------
    # 2. SÉLECTION DE LA LISTE DES ÉLÈVES (.xlsx ou .txt)
    # -------------------------------------------------------------------------
    print("---------------------------------------------------------------")
    print("📌 2. LISTE DES ÉLÈVES (.xlsx ou .txt)")
    fichiers_liste = glob.glob("*.xlsx") + glob.glob("*.txt")
    liste_path = ""

    if fichiers_liste:
        print("Fichiers trouvés dans le dossier actuel :")
        valid_opts_l = []
        def_idx_l = 1
        for idx, f in enumerate(fichiers_liste, 1):
            print(f"  [{idx}] {f}")
            valid_opts_l.append(str(idx))
            if "liste" in f.lower() or f == "eleves.txt":
                def_idx_l = idx

        if sys.platform == "darwin":
            print("  [M] Ouvrir le sélecteur macOS")
            valid_opts_l.append("M")
        print("  [S] Saisir le chemin manuellement")
        valid_opts_l.append("S")

        choix = get_valid_input(f"\n👉 Choix (défaut={def_idx_l}) : ", valid_opts_l, default=def_idx_l)
        
        if choix == 'M':
            liste_path = get_native_file_picker("Sélectionnez le fichier liste d'élèves")
        elif choix == 'S':
            while True:
                liste_path = input("Chemin du fichier liste : ").strip()
                if liste_path and os.path.exists(liste_path):
                    break
                print(f"❌ Fichier introuvable : '{liste_path}'. Réessayez.")
        else:
            liste_path = fichiers_liste[int(choix) - 1]
    else:
        if sys.platform == "darwin":
            liste_path = get_native_file_picker("Sélectionnez le fichier liste d'élèves")
        while not liste_path or not os.path.exists(liste_path):
            liste_path = input("👉 Entrez le chemin du fichier liste : ").strip()
            if not os.path.exists(liste_path):
                print(f"❌ Fichier introuvable : '{liste_path}'")

    eleves = parse_student_list(liste_path)
    while not eleves:
        print(f"\n❌ Impossible d'extraire des élèves depuis '{liste_path}'. Veuillez choisir un autre fichier.")
        liste_path = input("👉 Chemin du fichier liste : ").strip()
        if os.path.exists(liste_path):
            eleves = parse_student_list(liste_path)

    print(f"✅ Liste chargée : {len(eleves)} élève(s) trouvé(s)\n")

    # -------------------------------------------------------------------------
    # 3. DOSSIERS CIBLES DE DESTINATION (SIMPLE OU MULTIPLE)
    # -------------------------------------------------------------------------
    print("---------------------------------------------------------------")
    print("📌 3. DOSSIER(S) DE DESTINATION")
    print("  [1] Dossier unique par défaut ('./fichiers_eleves')")
    print("  [2] Sélection multiple de dossiers & Navigateur (Générer dans plusieurs dossiers)")
    if sys.platform == "darwin":
        print("  [M] Choisir un dossier avec le sélecteur macOS")

    opts_d = ["1", "2"]
    if sys.platform == "darwin":
        opts_d.append("M")

    choix_dossier = get_valid_input("\n👉 Choix (défaut=1) : ", opts_d, default="1")
    dossiers_cibles = []

    if choix_dossier == '2':
        dossiers_cibles = multi_folder_browser()
    elif choix_dossier == 'M':
        mac_folder = get_native_folder_picker("Sélectionnez le dossier de destination")
        if mac_folder:
            dossiers_cibles = [os.path.abspath(mac_folder)]
        else:
            dossiers_cibles = [os.path.abspath("fichiers_eleves")]
    else:
        nom_defaut = "fichiers_eleves"
        saisie = input(f"Nom du dossier à créer/utiliser (défaut='{nom_defaut}') : ").strip()
        dossiers_cibles = [os.path.abspath(saisie if saisie else nom_defaut)]

    for d in dossiers_cibles:
        os.makedirs(d, exist_ok=True)
    
    print(f"✅ {len(dossiers_cibles)} dossier(s) cible(s) prêt(s)\n")

    # -------------------------------------------------------------------------
    # 4. OPTIONS DE FORMATAGE DES NOMS ET CELLULE C3
    # -------------------------------------------------------------------------
    print("---------------------------------------------------------------")
    print("📌 4. FORMATAGE DU NOM DE FICHIER ET CELLULE C3")
    print("Format du nom de fichier :")
    print("  [1] Avec tiret bas '_'  -> 1M1_Nom_Prenom.xlsx (Recommandé)")
    print("  [2] Avec espaces ' '   -> 1M1 Nom Prenom.xlsx")
    fmt_choix = get_valid_input("👉 Choix (défaut=1) : ", ["1", "2"], default="1")
    separateur = " " if fmt_choix == "2" else "_"

    print("\nFormat de la cellule C3 :")
    print("  [1] Prénom Nom (ex: Alice Dupont)")
    print("  [2] Nom Prénom (ex: Dupont Alice)")
    c3_choix = get_valid_input("👉 Choix (défaut=1) : ", ["1", "2"], default="1")
    c3_format = "nom_prenom" if c3_choix == "2" else "prenom_nom"

    # -------------------------------------------------------------------------
    # 5. RÉCAPITULATIF ET CONFIRMATION
    # -------------------------------------------------------------------------
    print("\n===============================================================")
    print("📋 RÉCAPITULATIF DE L'AUTOMATISATION")
    print("===============================================================")
    print(f" 📄 Template       : {template_path}")
    print(f" 👥 Liste élèves   : {liste_path} ({len(eleves)} élèves)")
    print(f" 📁 Dossier(s) ({len(dossiers_cibles)}) :")
    for d in dossiers_cibles:
        print(f"     • {d}")
    print(f" ✏️ Cellule C3      : {'Nom Prénom' if c3_format == 'nom_prenom' else 'Prénom Nom'}")
    
    print("\n🔍 Aperçu des premiers fichiers qui vont être créés :")
    exemples = eleves[:3]
    for e in exemples:
        nom = e["nom"]
        prenom = e["prenom"]
        classe = e["classe"]
        parts = [p for p in [classe, nom, prenom] if p]
        nom_f = ("_".join(parts).replace(" ", "_") if separateur == "_" else " ".join(parts)) + ".xlsx"
        c3_val = f"{nom} {prenom}" if c3_format == "nom_prenom" else f"{prenom} {nom}"
        print(f"   • Nom de fichier : {nom_f} | C3 = '{c3_val.strip()}'")
    if len(eleves) > 3:
        print(f"   ... et {len(eleves) - 3} autre(s)")

    print("===============================================================")
    confirm = get_valid_input("\n🚀 Lancer la génération ? [O/N] : ", ["O", "N", "OUI", "NON", "Y", "YES"], default="O")

    if confirm in ['O', 'OUI', 'Y', 'YES']:
        print("\n⏳ Traitement en cours...")
        total_files = 0
        for target_dir in dossiers_cibles:
            print(f"\n📂 Enregistrement dans : {target_dir}")
            count = 0
            for e in eleves:
                nom = e["nom"]
                prenom = e["prenom"]
                classe = e["classe"]
                parts = [p for p in [classe, nom, prenom] if p]
                
                nom_f = ("_".join(parts).replace(" ", "_") if separateur == "_" else " ".join(parts)) + ".xlsx"
                c3_val = f"{nom} {prenom}" if c3_format == "nom_prenom" else f"{prenom} {nom}"

                destination = os.path.join(target_dir, nom_f)
                
                wb = openpyxl.load_workbook(template_path)
                ws = wb.active
                ws['C3'] = c3_val.strip()
                wb.save(destination)
                
                count += 1
                total_files += 1
                print(f"  [OK {count}/{len(eleves)}] {nom_f}")

        print(f"\n🎉 SUCCÈS ! {total_files} fichier(s) généré(s) à travers {len(dossiers_cibles)} dossier(s) !")
    else:
        print("\n❌ Opération annulée.")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nInterrompu par l'utilisateur.")
        sys.exit(0)
