#!/usr/bin/env python3
"""
Script de duplication automatique de templates Excel par élève.
- Extrait les listes d'élèves (fichiers .xlsx ou .txt)
- Remplit la cellule C3 du template Excel avec "Prénom Nom"
- Génère les fichiers au format "[Classe]_[Nom]_[Prénom].xlsx"
- Gère le TRAITEMENT PAR LOTS (plusieurs classes en 1 seule exécution)
"""

import os
import sys
import glob
import subprocess

def ensure_openpyxl():
    """Vérifie si openpyxl est installé et propose de l'installer si nécessaire."""
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("===============================================================")
        print(" ⚠️ DÉPENDANCE MANQUANTE : openpyxl")
        print("===============================================================")
        print("La bibliothèque 'openpyxl' est nécessaire pour lire et modifier")
        print("les fichiers Excel (.xlsx).\n")
        
        choix = input("👉 Souhaitez-vous l'installer automatiquement maintenant ? [O/n] : ").strip().lower()
        if choix in ['', 'o', 'oui', 'y', 'yes']:
            print("\n⏳ Installation de 'openpyxl' via pip...")
            try:
                subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            except Exception:
                try:
                    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages"])
                except Exception as err:
                    print(f"\n❌ Échec de l'installation automatique : {err}")
                    print("💡 Veuillez installer la bibliothèque manuellement avec :")
                    print("   pip3 install openpyxl")
                    sys.exit(1)
                    
            print("✅ 'openpyxl' a été installé avec succès !\n")
            import openpyxl
            return openpyxl
        else:
            print("\n❌ Impossible de continuer sans 'openpyxl'.")
            sys.exit(1)

openpyxl = ensure_openpyxl()

def find_files():
    """Détecte le template et tous les fichiers de listes de classes disponibles."""
    all_xlsx = glob.glob("*.xlsx")
    
    template_file = None
    listes_files = []
    
    for f in all_xlsx:
        f_lower = f.lower()
        if "liste" in f_lower:
            listes_files.append(f)
        elif "copie" in f_lower or "template" in f_lower:
            template_file = f
            
    if not template_file:
        for f in all_xlsx:
            if f not in listes_files:
                template_file = f
                break
                
    if not listes_files:
        for f in all_xlsx:
            if f != template_file:
                listes_files.append(f)
                
    if not listes_files and os.path.exists("eleves.txt"):
        listes_files.append("eleves.txt")
        
    return template_file, listes_files

def parse_student_list(liste_path):
    """Extrait les élèves depuis un fichier Excel (.xlsx) ou Texte (.txt)."""
    eleves = []
    
    if liste_path.endswith(".xlsx"):
        wb = openpyxl.load_workbook(liste_path, data_only=True)
        ws = wb.active
        
        header_row = None
        for r in range(1, 30):
            vals = [str(ws.cell(row=r, column=c).value or '').strip() for c in range(1, 15)]
            if 'Nom' in vals and ('Prénom' in vals or 'Prenom' in vals):
                header_row = r
                break
                
        if not header_row:
            return []
            
        nom_c = next(c for c in range(1, 15) if str(ws.cell(row=header_row, column=c).value).strip().lower() == 'nom')
        prenom_c = next(c for c in range(1, 15) if str(ws.cell(row=header_row, column=c).value).strip().lower() in ['prénom', 'prenom'])
        group_c = next((c for c in range(1, 15) if str(ws.cell(row=header_row, column=c).value).strip().lower() in ['groupe', 'classe']), None)
        
        for r in range(header_row + 1, ws.max_row + 1):
            nom = ws.cell(row=r, column=nom_c).value
            prenom = ws.cell(row=r, column=prenom_c).value
            grp = ws.cell(row=r, column=group_c).value if group_c else ''
            
            if nom and prenom:
                eleves.append({
                    "nom": str(nom).strip(),
                    "prenom": str(prenom).strip(),
                    "classe": str(grp).strip() if grp else ""
                })
    else:
        with open(liste_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    parts = line.split()
                    if len(parts) >= 2:
                        prenom = parts[0]
                        nom = " ".join(parts[1:])
                    else:
                        prenom = line
                        nom = ""
                    eleves.append({"nom": nom, "prenom": prenom, "classe": ""})
                    
    return eleves

def main():
    template_file, listes_files = find_files()
    
    if len(sys.argv) > 1:
        template_file = sys.argv[1]
    if len(sys.argv) > 2:
        listes_files = sys.argv[2:]
        
    print("===============================================================")
    print(" 🎓 DÉMARRAGE DU SCRIPT DE DUPLICATION (MULTI-CLASSES)")
    print("===============================================================")
    print(f" 📄 Template       : {template_file}")
    print(f" 👥 Listes classes : {len(listes_files)} fichier(s) détecté(s)")
    for lf in listes_files:
        print(f"     • {lf}")
    print("---------------------------------------------------------------")
    
    if not template_file or not os.path.exists(template_file):
        print("❌ Erreur : Fichier template introuvable.")
        sys.exit(1)
        
    if not listes_files:
        print("❌ Erreur : Aucun fichier de liste d'élèves trouvé.")
        sys.exit(1)

    parent_dir = "fichiers_eleves"
    os.makedirs(parent_dir, exist_ok=True)
    
    total_fichiers = 0
    
    for liste_file in listes_files:
        if not os.path.exists(liste_file):
            print(f"⚠️ Fichier introuvable : '{liste_file}'. Ignoré.")
            continue
            
        eleves = parse_student_list(liste_file)
        if not eleves:
            print(f"⚠️ Aucun élève extrait de '{liste_file}'. Ignoré.")
            continue
            
        nom_classe = eleves[0]["classe"] if eleves[0]["classe"] else os.path.splitext(os.path.basename(liste_file))[0]
        
        target_dir = os.path.join(parent_dir, nom_classe) if len(listes_files) > 1 else parent_dir
        os.makedirs(target_dir, exist_ok=True)
        
        print(f"\n📂 Classe '{nom_classe}' ({len(eleves)} élèves) -> Dossier: {os.path.abspath(target_dir)}")
        
        count = 0
        for e in eleves:
            nom = e["nom"]
            prenom = e["prenom"]
            classe = e["classe"] if e["classe"] else nom_classe
            
            parts = [p for p in [classe, nom, prenom] if p]
            nom_fichier = "_".join(parts).replace(" ", "_") + ".xlsx"
            
            wb = openpyxl.load_workbook(template_file)
            ws = wb.active
            
            ws['C3'] = f"{prenom} {nom}".strip()
            
            dest_path = os.path.join(target_dir, nom_fichier)
            wb.save(dest_path)
            count += 1
            total_fichiers += 1
            print(f"  ✅ [{count}/{len(eleves)}] Généré : {nom_fichier} (Cellule C3 = '{ws['C3'].value}')")
            
    print("\n===============================================================")
    print(f"🎉 SUCCÈS ! {total_fichiers} fichier(s) généré(s) pour {len(listes_files)} classe(s) !")
    print(f"📁 Localisation : {os.path.abspath(parent_dir)}")
    print("===============================================================")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nInterrompu par l'utilisateur.")
        sys.exit(0)
