#!/usr/bin/env python3
"""
🎓 Générateur de Fichiers Élèves - Interface Graphique (GUI)
Style Minimaliste Zinc & Violet Foncé.
- Gestionnaire de versioning & comparaison de Hash/Commit local vs GitHub
"""

import os
import sys
import glob
import subprocess
import json
import urllib.request
import zipfile
import io
import openpyxl

GITHUB_REPO = "Examera1005/list_to_excel"
GITHUB_API_URL = f"https://api.github.com/repos/{GITHUB_REPO}/commits/master"
GITHUB_ZIP_URL = f"https://github.com/{GITHUB_REPO}/archive/refs/heads/master.zip"

def get_local_commit():
    """Lit le hash de version/commit local enregistré."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    version_file = os.path.join(script_dir, ".version")
    if os.path.exists(version_file):
        try:
            with open(version_file, "r", encoding="utf-8") as f:
                val = f.read().strip()
                if val:
                    return val
        except Exception:
            pass
    try:
        res = subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], cwd=script_dir, stderr=subprocess.DEVNULL).decode().strip()
        if res:
            return res
    except Exception:
        pass
    return "v1.0.0"

def save_local_commit(commit_sha):
    """Enregistre le hash du commit localement après une mise à jour réussie."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    version_file = os.path.join(script_dir, ".version")
    try:
        with open(version_file, "w", encoding="utf-8") as f:
            f.write(commit_sha)
    except Exception:
        pass

def check_github_update():
    """Vérifie le dernier commit sur GitHub master."""
    try:
        req = urllib.request.Request(GITHUB_API_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=4) as response:
            data = json.loads(response.read().decode())
            return data['sha'][:7]
    except Exception:
        return None

def download_and_apply_update():
    """Télécharge et extrait automatiquement les derniers fichiers depuis GitHub."""
    try:
        req = urllib.request.Request(GITHUB_ZIP_URL, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            zip_bytes = response.read()

        script_dir = os.path.dirname(os.path.abspath(__file__))
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            for member in z.namelist():
                filename = os.path.basename(member)
                if filename and not filename.startswith('.'):
                    if filename in ['app_gui.py', 'app_tui.py', 'dupliquer.py', 'build_mac_app.py', 'README.md', 'template.xlsx', 'liste_eleves_1M1.xlsx', 'liste_eleves_1M2.xlsx']:
                        dest_path = os.path.join(script_dir, filename)
                        with z.open(member) as src_f, open(dest_path, 'wb') as dst_f:
                            dst_f.write(src_f.read())
        return True
    except Exception:
        return False

def parse_student_list(liste_path):
    """Extrait les élèves depuis un fichier Excel (.xlsx) ou Texte (.txt)."""
    eleves = []
    if liste_path.endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(liste_path, data_only=True)
            ws = wb.active
            header_row = None
            for r in range(1, 30):
                vals = [str(ws.cell(row=r, column=c).value or '').strip() for c in range(1, 15)]
                if 'Nom' in vals and ('Prénom' in vals or 'Prenom' in vals):
                    header_row = r
                    break
            if header_row:
                nom_c = next(c for c in range(1, 15) if str(ws.cell(row=header_row, column=c).value).strip().lower() == 'nom')
                prenom_c = next(c for c in range(1, 15) if str(ws.cell(row=header_row, column=c).value).strip().lower() in ['prénom', 'prenom'])
                group_c = next((c for c in range(1, 15) if str(ws.cell(row=header_row, column=c).value).strip().lower() in ['groupe', 'classe']), None)
                for r in range(header_row + 1, ws.max_row + 1):
                    nom = ws.cell(row=r, column=nom_c).value
                    prenom = ws.cell(row=r, column=prenom_c).value
                    grp = ws.cell(row=r, column=group_c).value if group_c else ''
                    if nom and prenom:
                        eleves.append({"nom": str(nom).strip(), "prenom": str(prenom).strip(), "classe": str(grp).strip() if grp else ""})
                return eleves
        except Exception:
            pass

    try:
        with open(liste_path, "r", encoding="utf-8", errors="ignore") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):
                    parts = line.split()
                    if len(parts) >= 2:
                        prenom = parts[0]
                        nom = " ".join(parts[1:])
                    else:
                        prenom = line
                        nom = ""
                    eleves.append({"nom": nom, "prenom": prenom, "classe": ""})
    except Exception:
        pass
    return eleves


# =============================================================================
# MOTEUR 1 : PySide6 (Qt6) - Design Zinc & Violet Minimaliste
# =============================================================================
def run_pyside6_app():
    from PySide6 import QtWidgets, QtCore, QtGui

    class WorkerThread(QtCore.QThread):
        progress_signal = QtCore.Signal(int, int, str)
        finished_signal = QtCore.Signal(int, int, str)
        error_signal = QtCore.Signal(str)

        def __init__(self, template_path, listes_files, parent_dir, separateur, c3_format):
            super().__init__()
            self.template_path = template_path
            self.listes_files = listes_files
            self.parent_dir = parent_dir
            self.separateur = separateur
            self.c3_format = c3_format

        def run(self):
            try:
                total_fichiers = 0
                all_data = []
                for l_file in self.listes_files:
                    eleves = parse_student_list(l_file)
                    if eleves:
                        all_data.append((l_file, eleves))

                total_items = sum(len(e) for _, e in all_data)
                current_item = 0

                for l_file, eleves in all_data:
                    nom_classe = eleves[0]["classe"] if eleves[0]["classe"] else os.path.splitext(os.path.basename(l_file))[0]
                    target_dir = os.path.join(self.parent_dir, nom_classe) if len(all_data) > 1 else self.parent_dir
                    os.makedirs(target_dir, exist_ok=True)

                    for e in eleves:
                        nom = e["nom"]
                        prenom = e["prenom"]
                        classe = e["classe"] if e["classe"] else nom_classe
                        parts = [p for p in [classe, nom, prenom] if p]
                        
                        nom_f = ("_".join(parts).replace(" ", "_") if self.separateur == "_" else " ".join(parts)) + ".xlsx"
                        c3_val = f"{nom} {prenom}" if self.c3_format == "nom_prenom" else f"{prenom} {nom}"

                        destination = os.path.join(target_dir, nom_f)
                        
                        wb = openpyxl.load_workbook(self.template_path)
                        ws = wb.active
                        ws['C3'] = c3_val.strip()
                        wb.save(destination)

                        current_item += 1
                        total_fichiers += 1
                        msg = f"✓ [{current_item}/{total_items}] [{nom_classe}] Généré : {nom_f}"
                        self.progress_signal.emit(current_item, total_items, msg)

                self.finished_signal.emit(total_fichiers, len(all_data), os.path.abspath(self.parent_dir))
            except Exception as err:
                self.error_signal.emit(str(err))

    class MainWindow(QtWidgets.QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Générateur de Documents Élèves")
            self.setMinimumSize(880, 720)
            self.resize(920, 760)

            self.template_path = ""
            self.listes_files = []
            self.parent_dir = os.path.abspath("fichiers_eleves")

            self.init_ui()
            self.auto_detect_files()

        def init_ui(self):
            self.setStyleSheet("""
                QMainWindow, QDialog, QMessageBox {
                    background-color: #09090B;
                    color: #FAFAFA;
                }
                QWidget {
                    color: #FAFAFA;
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
                    font-size: 13px;
                }
                QGroupBox {
                    background-color: #18181B;
                    border: 1px solid #27272A;
                    border-radius: 10px;
                    margin-top: 12px;
                    padding: 16px;
                    font-weight: 600;
                    font-size: 13px;
                    color: #A78BFA;
                }
                QGroupBox::title {
                    subcontrol-origin: margin;
                    left: 12px;
                    padding: 0 4px;
                }
                QPushButton {
                    background-color: #27272A;
                    color: #FAFAFA;
                    border: 1px solid #3F3F46;
                    border-radius: 6px;
                    padding: 8px 14px;
                    font-weight: 500;
                }
                QPushButton:hover {
                    background-color: #3F3F46;
                    border-color: #52525B;
                }
                QPushButton:pressed {
                    background-color: #52525B;
                }
                QPushButton#btnLaunch {
                    background-color: #7C3AED;
                    color: #FFFFFF;
                    border: none;
                    font-size: 14px;
                    padding: 12px 20px;
                    font-weight: 600;
                    border-radius: 8px;
                }
                QPushButton#btnLaunch:hover {
                    background-color: #6D28D9;
                }
                QPushButton#btnLaunch:pressed {
                    background-color: #5B21B6;
                }
                QLineEdit {
                    background-color: #18181B;
                    border: 1px solid #3F3F46;
                    border-radius: 6px;
                    padding: 8px 12px;
                    color: #FAFAFA;
                    selection-background-color: #7C3AED;
                }
                QLineEdit:focus {
                    border-color: #7C3AED;
                }
                QListWidget {
                    background-color: #18181B;
                    border: 1px solid #3F3F46;
                    border-radius: 6px;
                    padding: 4px;
                    color: #FAFAFA;
                    selection-background-color: #7C3AED;
                    selection-color: #FFFFFF;
                }
                QListWidget::item {
                    padding: 6px 10px;
                    border-radius: 4px;
                    color: #FAFAFA;
                }
                QListWidget::item:hover {
                    background-color: #27272A;
                }
                QRadioButton {
                    color: #D4D4D8;
                    spacing: 8px;
                }
                QRadioButton::indicator {
                    width: 16px;
                    height: 16px;
                    border-radius: 8px;
                    border: 1px solid #52525B;
                    background-color: #18181B;
                }
                QRadioButton::indicator:checked {
                    background-color: #7C3AED;
                    border-color: #A78BFA;
                }
                QProgressBar {
                    background-color: #18181B;
                    border: 1px solid #3F3F46;
                    border-radius: 6px;
                    text-align: center;
                    color: #FAFAFA;
                    font-weight: 600;
                }
                QProgressBar::chunk {
                    background-color: #7C3AED;
                    border-radius: 5px;
                }
                QTextEdit {
                    background-color: #18181B;
                    border: 1px solid #3F3F46;
                    border-radius: 6px;
                    color: #A78BFA;
                    font-family: monospace;
                    font-size: 12px;
                    padding: 8px;
                }
                QMessageBox {
                    background-color: #18181B;
                    color: #FAFAFA;
                    border: 1px solid #3F3F46;
                }
                QMessageBox QLabel {
                    color: #FAFAFA;
                    background-color: transparent;
                }
                QMessageBox QPushButton {
                    background-color: #27272A;
                    color: #FAFAFA;
                    border: 1px solid #3F3F46;
                    border-radius: 6px;
                    padding: 6px 16px;
                    min-width: 70px;
                }
            """)

            self.completer_model = QtWidgets.QFileSystemModel(self)
            self.completer_model.setRootPath("")

            central_widget = QtWidgets.QWidget()
            self.setCentralWidget(central_widget)
            main_layout = QtWidgets.QVBoxLayout(central_widget)
            main_layout.setContentsMargins(24, 24, 24, 24)
            main_layout.setSpacing(14)

            # En-tête avec bouton de mise à jour GitHub
            header_layout = QtWidgets.QHBoxLayout()
            header_text_layout = QtWidgets.QVBoxLayout()
            title_label = QtWidgets.QLabel("Générateur de Documents Élèves")
            title_label.setStyleSheet("font-size: 18px; font-weight: 700; color: #FAFAFA;")
            
            current_local = get_local_commit()
            subtitle_label = QtWidgets.QLabel(f"Duplication automatique par classe (Version : {current_local})")
            subtitle_label.setStyleSheet("font-size: 12px; color: #A1A1AA;")
            header_text_layout.addWidget(title_label)
            header_text_layout.addWidget(subtitle_label)
            
            btn_update = QtWidgets.QPushButton("🔄 Vérifier mise à jour GitHub")
            btn_update.setStyleSheet("background-color: #18181B; color: #A78BFA; border-color: #7C3AED;")
            btn_update.setCursor(QtCore.Qt.PointingHandCursor)
            btn_update.clicked.connect(self.check_and_update_gui)

            header_layout.addLayout(header_text_layout)
            header_layout.addStretch()
            header_layout.addWidget(btn_update)
            main_layout.addLayout(header_layout)

            # 1. Template
            group_template = QtWidgets.QGroupBox("1. Fichier Modèle Template Excel (.xlsx)")
            gt_layout = QtWidgets.QHBoxLayout(group_template)
            self.txt_template = QtWidgets.QLineEdit()
            self.txt_template.setPlaceholderText("Saisissez un chemin (Tab zsh/fish) ou cliquez sur Parcourir...")
            
            completer_tpl = QtWidgets.QCompleter(self.completer_model, self)
            completer_tpl.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            self.txt_template.setCompleter(completer_tpl)
            self.txt_template.textChanged.connect(self.on_template_text_changed)

            btn_browse_tpl = QtWidgets.QPushButton("Parcourir...")
            btn_browse_tpl.setCursor(QtCore.Qt.PointingHandCursor)
            btn_browse_tpl.clicked.connect(self.browse_template)
            gt_layout.addWidget(self.txt_template)
            gt_layout.addWidget(btn_browse_tpl)
            main_layout.addWidget(group_template)

            # 2. Listes
            group_listes = QtWidgets.QGroupBox("2. Listes de Classes (.xlsx ou .txt)")
            gl_layout = QtWidgets.QVBoxLayout(group_listes)
            
            gl_top = QtWidgets.QHBoxLayout()
            self.lbl_listes_count = QtWidgets.QLabel("0 classe(s) sélectionnée(s)")
            self.lbl_listes_count.setStyleSheet("color: #A1A1AA; font-weight: normal;")
            
            self.txt_filter_fzf = QtWidgets.QLineEdit()
            self.txt_filter_fzf.setPlaceholderText("🔍 Filtrer les listes (fzf)...")
            self.txt_filter_fzf.setFixedWidth(200)
            self.txt_filter_fzf.textChanged.connect(self.filter_listes_fzf)

            btn_browse_listes = QtWidgets.QPushButton("Ajouter des listes (Finder / Explorateur)...")
            btn_browse_listes.setCursor(QtCore.Qt.PointingHandCursor)
            btn_browse_listes.clicked.connect(self.browse_listes)
            
            btn_remove_selected = QtWidgets.QPushButton("Supprimer sélection")
            btn_remove_selected.setStyleSheet("background-color: #27272A; color: #F59E0B; border-color: #D97706;")
            btn_remove_selected.setCursor(QtCore.Qt.PointingHandCursor)
            btn_remove_selected.clicked.connect(self.remove_selected_listes)

            btn_clear_listes = QtWidgets.QPushButton("Tout effacer")
            btn_clear_listes.setStyleSheet("background-color: #27272A; color: #F43F5E; border-color: #E11D48;")
            btn_clear_listes.setCursor(QtCore.Qt.PointingHandCursor)
            btn_clear_listes.clicked.connect(self.clear_listes)

            gl_top.addWidget(self.lbl_listes_count)
            gl_top.addWidget(self.txt_filter_fzf)
            gl_top.addStretch()
            gl_top.addWidget(btn_browse_listes)
            gl_top.addWidget(btn_remove_selected)
            gl_top.addWidget(btn_clear_listes)
            
            self.lst_listes = QtWidgets.QListWidget()
            self.lst_listes.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
            self.lst_listes.setMaximumHeight(100)
            gl_layout.addLayout(gl_top)
            gl_layout.addWidget(self.lst_listes)
            main_layout.addWidget(group_listes)

            # 3. Destination
            group_dest = QtWidgets.QGroupBox("3. Dossier de Destination Parent")
            gd_layout = QtWidgets.QHBoxLayout(group_dest)
            self.txt_dest = QtWidgets.QLineEdit(self.parent_dir)
            
            completer_dest = QtWidgets.QCompleter(self.completer_model, self)
            completer_dest.setCaseSensitivity(QtCore.Qt.CaseInsensitive)
            self.txt_dest.setCompleter(completer_dest)

            btn_browse_dest = QtWidgets.QPushButton("Choisir le dossier...")
            btn_browse_dest.setCursor(QtCore.Qt.PointingHandCursor)
            btn_browse_dest.clicked.connect(self.browse_dest)
            gd_layout.addWidget(self.txt_dest)
            gd_layout.addWidget(btn_browse_dest)
            main_layout.addWidget(group_dest)

            # 4. Options
            group_opts = QtWidgets.QGroupBox("4. Options de Nommage & Cellule C3")
            go_layout = QtWidgets.QHBoxLayout(group_opts)
            sep_layout = QtWidgets.QVBoxLayout()
            sep_title = QtWidgets.QLabel("Format du nom de fichier :")
            sep_title.setStyleSheet("color: #E4E4E7; font-weight: 600;")
            self.rad_sep_underscore = QtWidgets.QRadioButton("Avec tirets bas '_' (1M1_Nom_Prenom.xlsx)")
            self.rad_sep_underscore.setChecked(True)
            self.rad_sep_space = QtWidgets.QRadioButton("Avec espaces ' ' (1M1 Nom Prenom.xlsx)")
            sep_layout.addWidget(sep_title)
            sep_layout.addWidget(self.rad_sep_underscore)
            sep_layout.addWidget(self.rad_sep_space)
            
            c3_layout = QtWidgets.QVBoxLayout()
            c3_title = QtWidgets.QLabel("Format de la cellule C3 :")
            c3_title.setStyleSheet("color: #E4E4E7; font-weight: 600;")
            self.rad_c3_prenom_nom = QtWidgets.QRadioButton("Prénom Nom (ex: Alice Dupont)")
            self.rad_c3_prenom_nom.setChecked(True)
            self.rad_c3_nom_prenom = QtWidgets.QRadioButton("Nom Prénom (ex: Dupont Alice)")
            c3_layout.addWidget(c3_title)
            c3_layout.addWidget(self.rad_c3_prenom_nom)
            c3_layout.addWidget(self.rad_c3_nom_prenom)

            go_layout.addLayout(sep_layout)
            go_layout.addSpacing(30)
            go_layout.addLayout(c3_layout)
            main_layout.addWidget(group_opts)

            # Lancement & Progression
            self.btn_launch = QtWidgets.QPushButton("Lancer la génération par lots")
            self.btn_launch.setObjectName("btnLaunch")
            self.btn_launch.setCursor(QtCore.Qt.PointingHandCursor)
            self.btn_launch.clicked.connect(self.start_generation)
            main_layout.addWidget(self.btn_launch)

            self.progress_bar = QtWidgets.QProgressBar()
            self.progress_bar.setVisible(False)
            main_layout.addWidget(self.progress_bar)

            self.txt_log = QtWidgets.QTextEdit()
            self.txt_log.setReadOnly(True)
            self.txt_log.setMaximumHeight(110)
            main_layout.addWidget(self.txt_log)

        def check_and_update_gui(self):
            local_sha = get_local_commit()
            latest_sha = check_github_update()

            if not latest_sha:
                self.show_styled_dialog(QtWidgets.QMessageBox.Warning, "GitHub Status", "Impossible de contacter GitHub pour vérifier les mises à jour.")
                return

            if local_sha == latest_sha:
                self.show_styled_dialog(
                    QtWidgets.QMessageBox.Information,
                    "Application à jour",
                    f"✅ Votre application est déjà à jour !\n\n• Version locale : {local_sha}\n• Version GitHub : {latest_sha}"
                )
                return

            reply = QtWidgets.QMessageBox.question(
                self, "Mise à jour GitHub disponible",
                f"🎉 Une nouvelle version est disponible sur GitHub !\n\n• Version actuelle : {local_sha}\n• Nouvelle version : {latest_sha}\n\nVoulez-vous la télécharger ?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
            )
            if reply == QtWidgets.QMessageBox.Yes:
                if download_and_apply_update():
                    save_local_commit(latest_sha)
                    self.show_styled_dialog(
                        QtWidgets.QMessageBox.Information,
                        "Mise à jour réussie !",
                        f"✅ L'application a été mise à jour avec succès vers la version {latest_sha} !\nVeuillez redémarrer l'application."
                    )
                else:
                    self.show_styled_dialog(QtWidgets.QMessageBox.Warning, "Échec", "La mise à jour automatique a échoué.")

        def auto_detect_files(self):
            all_xlsx = glob.glob("*.xlsx")
            vrais_templates = []
            for f in all_xlsx:
                try:
                    wb = openpyxl.load_workbook(f, read_only=True)
                    wb.close()
                    vrais_templates.append(os.path.abspath(f))
                except Exception:
                    pass
            if vrais_templates:
                self.template_path = vrais_templates[0]
                self.txt_template.setText(self.template_path)
            listes = [os.path.abspath(f) for f in glob.glob("*.xlsx") + glob.glob("*.txt") if os.path.abspath(f) != self.template_path]
            if listes:
                self.listes_files = listes
                self.update_listes_widget()

        def on_template_text_changed(self, text):
            if os.path.exists(text.strip()):
                self.template_path = text.strip()

        def browse_template(self):
            filePath, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Sélectionnez le template Excel", "", "Fichiers Excel (*.xlsx)")
            if filePath:
                self.template_path = filePath
                self.txt_template.setText(self.template_path)

        def browse_listes(self):
            files, _ = QtWidgets.QFileDialog.getOpenFileNames(self, "Sélectionnez les listes d'élèves", "", "Listes d'élèves (*.xlsx *.txt)")
            if files:
                for f in files:
                    abs_f = os.path.abspath(f)
                    if abs_f not in self.listes_files:
                        self.listes_files.append(abs_f)
                self.update_listes_widget()

        def remove_selected_listes(self):
            selected_items = self.lst_listes.selectedItems()
            if not selected_items:
                return
            for item in selected_items:
                path = item.data(QtCore.Qt.UserRole)
                if path in self.listes_files:
                    self.listes_files.remove(path)
            self.update_listes_widget()

        def clear_listes(self):
            self.listes_files.clear()
            self.update_listes_widget()

        def filter_listes_fzf(self, query):
            query = query.strip().lower()
            self.lst_listes.clear()
            for f in self.listes_files:
                if not query or query in f.lower():
                    item = QtWidgets.QListWidgetItem(f"📄 {f}")
                    item.setData(QtCore.Qt.UserRole, f)
                    self.lst_listes.addItem(item)

        def update_listes_widget(self):
            self.filter_listes_fzf(self.txt_filter_fzf.text())
            self.lbl_listes_count.setText(f"{len(self.listes_files)} classe(s) sélectionnée(s)")

        def browse_dest(self):
            dirPath = QtWidgets.QFileDialog.getExistingDirectory(self, "Sélectionnez le dossier parent")
            if dirPath:
                self.parent_dir = dirPath
                self.txt_dest.setText(self.parent_dir)

        def show_styled_dialog(self, icon, title, text):
            msg = QtWidgets.QMessageBox(self)
            msg.setIcon(icon)
            msg.setWindowTitle(title)
            msg.setText(text)
            msg.exec()

        def start_generation(self):
            tpl = self.txt_template.text().strip()
            if tpl and os.path.exists(tpl):
                self.template_path = tpl

            if not self.template_path or not os.path.exists(self.template_path):
                self.show_styled_dialog(QtWidgets.QMessageBox.Warning, "Template manquant", "Veuillez sélectionner un fichier template Excel.")
                return
            if not self.listes_files:
                self.show_styled_dialog(QtWidgets.QMessageBox.Warning, "Listes manquantes", "Veuillez sélectionner au moins une liste d'élèves.")
                return
            parent_d = self.txt_dest.text().strip() or os.path.abspath("fichiers_eleves")
            separateur = "_" if self.rad_sep_underscore.isChecked() else " "
            c3_format = "nom_prenom" if self.rad_c3_nom_prenom.isChecked() else "prenom_nom"

            self.btn_launch.setEnabled(False)
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.txt_log.clear()
            self.txt_log.append("Démarrage de la génération par lots...\n")

            self.worker = WorkerThread(self.template_path, self.listes_files, parent_d, separateur, c3_format)
            self.worker.progress_signal.connect(self.on_progress)
            self.worker.finished_signal.connect(self.on_finished)
            self.worker.error_signal.connect(self.on_error)
            self.worker.start()

        def on_progress(self, current, total, message):
            self.progress_bar.setValue(int((current / total) * 100) if total > 0 else 0)
            self.txt_log.append(message)

        def on_finished(self, total_files, total_classes, parent_dir):
            self.progress_bar.setValue(100)
            self.btn_launch.setEnabled(True)
            self.txt_log.append(f"\n✓ Succès : {total_files} fichier(s) généré(s) pour {total_classes} classe(s) !")
            self.show_styled_dialog(QtWidgets.QMessageBox.Information, "Succès !", f"🎉 {total_files} fichier(s) généré(s) pour {total_classes} classe(s) !\n\n📁 Dossier : {parent_dir}")

        def on_error(self, err_msg):
            self.btn_launch.setEnabled(True)
            self.progress_bar.setVisible(False)
            self.show_styled_dialog(QtWidgets.QMessageBox.Critical, "Erreur", f"Erreur :\n{err_msg}")

    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


# =============================================================================
# MOTEUR 2 : Tkinter (Fallback Natif) - Design Zinc Minimaliste & Haut Contraste
# =============================================================================
def run_tkinter_app():
    import tkinter as tk
    from tkinter import filedialog, messagebox
    import threading

    class TkApp(tk.Tk):
        def __init__(self):
            super().__init__()
            self.title("Générateur de Documents Élèves")
            self.geometry("880x720")
            self.configure(bg="#09090B")

            self.template_path = ""
            self.listes_files = []
            self.parent_dir = os.path.abspath("fichiers_eleves")

            self.init_ui()
            self.auto_detect_files()

        def make_hover_button(self, parent, text, command, bg="#27272A", hover_bg="#3F3F46", fg="#FAFAFA", **kwargs):
            btn = tk.Button(parent, text=text, command=command, bg=bg, fg=fg, activebackground=hover_bg,
                            activeforeground="#FFFFFF", bd=1, relief="solid", highlightthickness=0,
                            font=("Segoe UI", 9, "bold"), cursor="hand2", **kwargs)
            btn.bind("<Enter>", lambda e: btn.config(bg=hover_bg))
            btn.bind("<Leave>", lambda e: btn.config(bg=bg))
            return btn

        def init_ui(self):
            main_frame = tk.Frame(self, bg="#09090B", padx=24, pady=24)
            main_frame.pack(fill="both", expand=True)

            # Header
            header_frame = tk.Frame(main_frame, bg="#09090B")
            header_frame.pack(fill="x", pady=(0, 16))
            lbl_title = tk.Label(header_frame, text="Générateur de Documents Élèves", font=("Segoe UI", 16, "bold"), bg="#09090B", fg="#FAFAFA")
            lbl_title.pack(side="left")
            
            btn_upd = self.make_hover_button(header_frame, "🔄 Mise à jour GitHub", self.check_and_update_tk, bg="#18181B", hover_bg="#27272A", fg="#A78BFA", padx=8, pady=3)
            btn_upd.pack(side="right")

            # 1. Template Group
            f1 = tk.LabelFrame(main_frame, text=" 1. Fichier Modèle Template Excel (.xlsx) ", font=("Segoe UI", 10, "bold"), bg="#18181B", fg="#A78BFA", padx=14, pady=10, bd=1, relief="solid")
            f1.pack(fill="x", pady=6)
            self.entry_tpl = tk.Entry(f1, bg="#18181B", fg="#FAFAFA", insertbackground="#FAFAFA", selectbackground="#7C3AED", selectforeground="#FFFFFF", bd=1, relief="solid", font=("Segoe UI", 10))
            self.entry_tpl.pack(side="left", fill="x", expand=True, padx=(0, 10), ipady=4)
            btn_tpl = self.make_hover_button(f1, "Parcourir...", self.browse_template, padx=12, pady=3)
            btn_tpl.pack(side="right")

            # 2. Listes Group
            f2 = tk.LabelFrame(main_frame, text=" 2. Listes de Classes (.xlsx ou .txt) ", font=("Segoe UI", 10, "bold"), bg="#18181B", fg="#A78BFA", padx=14, pady=10, bd=1, relief="solid")
            f2.pack(fill="x", pady=6)
            top_l = tk.Frame(f2, bg="#18181B")
            top_l.pack(fill="x", pady=(0, 6))
            self.lbl_count = tk.Label(top_l, text="0 classe(s) sélectionnée(s)", font=("Segoe UI", 9), bg="#18181B", fg="#A1A1AA")
            self.lbl_count.pack(side="left")

            self.entry_filter = tk.Entry(top_l, bg="#18181B", fg="#FAFAFA", insertbackground="#FAFAFA", bd=1, relief="solid", font=("Segoe UI", 9), width=22)
            self.entry_filter.pack(side="left", padx=(15, 0))
            self.entry_filter.bind("<KeyRelease>", self.filter_listes_fzf_tk)
            lbl_fzf = tk.Label(top_l, text="🔍 fzf", font=("Segoe UI", 9), bg="#18181B", fg="#71717A")
            lbl_fzf.pack(side="left", padx=(4, 0))

            btn_clear = self.make_hover_button(top_l, "Tout effacer", self.clear_listes, bg="#27272A", hover_bg="#3F3F46", fg="#F43F5E", padx=8, pady=2)
            btn_clear.pack(side="right", padx=(6, 0))
            btn_remove_sel = self.make_hover_button(top_l, "Supprimer sélection", self.remove_selected_listes, bg="#27272A", hover_bg="#3F3F46", fg="#F59E0B", padx=8, pady=2)
            btn_remove_sel.pack(side="right", padx=(6, 0))
            btn_listes = self.make_hover_button(top_l, "Ajouter des listes...", self.browse_listes, padx=12, pady=2)
            btn_listes.pack(side="right")
            
            self.listbox = tk.Listbox(f2, bg="#18181B", fg="#FAFAFA", selectmode="extended", selectbackground="#7C3AED", selectforeground="#FFFFFF", height=4, bd=1, relief="solid", font=("Segoe UI", 9))
            self.listbox.pack(fill="x")

            # 3. Destination Group
            f3 = tk.LabelFrame(main_frame, text=" 3. Dossier de Destination Parent ", font=("Segoe UI", 10, "bold"), bg="#18181B", fg="#A78BFA", padx=14, pady=10, bd=1, relief="solid")
            f3.pack(fill="x", pady=6)
            self.entry_dest = tk.Entry(f3, bg="#18181B", fg="#FAFAFA", insertbackground="#FAFAFA", selectbackground="#7C3AED", selectforeground="#FFFFFF", bd=1, relief="solid", font=("Segoe UI", 10))
            self.entry_dest.insert(0, self.parent_dir)
            self.entry_dest.pack(side="left", fill="x", expand=True, padx=(0, 10), ipady=4)
            btn_dest = self.make_hover_button(f3, "Choisir...", self.browse_dest, padx=12, pady=3)
            btn_dest.pack(side="right")

            # 4. Options Group
            f4 = tk.LabelFrame(main_frame, text=" 4. Options de Nommage & Cellule C3 ", font=("Segoe UI", 10, "bold"), bg="#18181B", fg="#A78BFA", padx=14, pady=10, bd=1, relief="solid")
            f4.pack(fill="x", pady=6)
            
            self.var_sep = tk.StringVar(value="_")
            lbl_sep = tk.Label(f4, text="Nom de fichier :", bg="#18181B", fg="#E4E4E7", font=("Segoe UI", 9, "bold"))
            lbl_sep.grid(row=0, column=0, sticky="w", padx=5)
            r1 = tk.Radiobutton(f4, text="Avec tiret '_' (1M1_Nom_Prenom.xlsx)", variable=self.var_sep, value="_", bg="#18181B", fg="#D4D4D8", selectcolor="#09090B", activebackground="#18181B")
            r1.grid(row=1, column=0, sticky="w", padx=5)
            r2 = tk.Radiobutton(f4, text="Avec espace ' ' (1M1 Nom Prenom.xlsx)", variable=self.var_sep, value=" ", bg="#18181B", fg="#D4D4D8", selectcolor="#09090B", activebackground="#18181B")
            r2.grid(row=2, column=0, sticky="w", padx=5)

            self.var_c3 = tk.StringVar(value="prenom_nom")
            lbl_c3 = tk.Label(f4, text="Format cellule C3 :", bg="#18181B", fg="#E4E4E7", font=("Segoe UI", 9, "bold"))
            lbl_c3.grid(row=0, column=1, sticky="w", padx=(30, 5))
            r3 = tk.Radiobutton(f4, text="Prénom Nom (ex: Alice Dupont)", variable=self.var_c3, value="prenom_nom", bg="#18181B", fg="#D4D4D8", selectcolor="#09090B", activebackground="#18181B")
            r3.grid(row=1, column=1, sticky="w", padx=(30, 5))
            r4 = tk.Radiobutton(f4, text="Nom Prénom (ex: Dupont Alice)", variable=self.var_c3, value="nom_prenom", bg="#18181B", fg="#D4D4D8", selectcolor="#09090B", activebackground="#18181B")
            r4.grid(row=2, column=1, sticky="w", padx=(30, 5))

            # Launch Button
            self.btn_launch = self.make_hover_button(main_frame, "Lancer la génération par lots", self.start_generation, bg="#7C3AED", hover_bg="#6D28D9", fg="#FFFFFF", pady=10)
            self.btn_launch.pack(fill="x", pady=12)

            # Console Log
            self.txt_log = tk.Text(main_frame, bg="#18181B", fg="#A78BFA", height=6, font=("Consolas", 9), bd=1, relief="solid")
            self.txt_log.pack(fill="both", expand=True)

        def check_and_update_tk(self):
            local_sha = get_local_commit()
            latest_sha = check_github_update()
            if not latest_sha:
                messagebox.showinfo("Mise à jour GitHub", "Impossible de contacter GitHub pour vérifier les mises à jour.")
                return
            if local_sha == latest_sha:
                messagebox.showinfo("Application à jour", f"✅ Votre application est déjà à jour !\n\n• Version locale : {local_sha}\n• Version GitHub : {latest_sha}")
                return
            if messagebox.askyesno("Mise à jour GitHub", f"Une nouvelle version est disponible sur GitHub !\n\n• Version actuelle : {local_sha}\n• Nouvelle version : {latest_sha}\n\nSouhaitez-vous mettre à jour les fichiers ?"):
                if download_and_apply_update():
                    save_local_commit(latest_sha)
                    messagebox.showinfo("Succès", f"L'application a été mise à jour vers la version {latest_sha} !")
                else:
                    messagebox.showerror("Échec", "La mise à jour a échoué.")

        def auto_detect_files(self):
            all_xlsx = glob.glob("*.xlsx")
            vrais_templates = []
            for f in all_xlsx:
                try:
                    wb = openpyxl.load_workbook(f, read_only=True)
                    wb.close()
                    vrais_templates.append(os.path.abspath(f))
                except Exception:
                    pass
            if vrais_templates:
                self.template_path = vrais_templates[0]
                self.entry_tpl.delete(0, tk.END)
                self.entry_tpl.insert(0, self.template_path)
            listes = [os.path.abspath(f) for f in glob.glob("*.xlsx") + glob.glob("*.txt") if os.path.abspath(f) != self.template_path]
            if listes:
                self.listes_files = listes
                self.update_listbox()

        def browse_template(self):
            path = filedialog.askopenfilename(title="Sélectionner le template Excel", filetypes=[("Excel Files", "*.xlsx")])
            if path:
                self.template_path = path
                self.entry_tpl.delete(0, tk.END)
                self.entry_tpl.insert(0, self.template_path)

        def browse_listes(self):
            files = filedialog.askopenfilenames(title="Sélectionner les listes d'élèves", filetypes=[("Listes Élèves", "*.xlsx *.txt")])
            if files:
                for f in files:
                    abs_f = os.path.abspath(f)
                    if abs_f not in self.listes_files:
                        self.listes_files.append(abs_f)
                self.update_listbox()

        def remove_selected_listes(self):
            sel = self.listbox.curselection()
            if not sel:
                return
            for idx in reversed(sel):
                item_text = self.listbox.get(idx).replace("📄 ", "")
                if item_text in self.listes_files:
                    self.listes_files.remove(item_text)
            self.update_listbox()

        def clear_listes(self):
            self.listes_files.clear()
            self.update_listbox()

        def filter_listes_fzf_tk(self, event=None):
            query = self.entry_filter.get().strip().lower()
            self.listbox.delete(0, tk.END)
            for f in self.listes_files:
                if not query or query in f.lower():
                    self.listbox.insert(tk.END, f"📄 {f}")

        def update_listbox(self):
            self.filter_listes_fzf_tk()
            self.lbl_count.config(text=f"{len(self.listes_files)} classe(s) sélectionnée(s)")

        def browse_dest(self):
            path = filedialog.askdirectory(title="Sélectionner le dossier parent")
            if path:
                self.parent_dir = path
                self.entry_dest.delete(0, tk.END)
                self.entry_dest.insert(0, self.parent_dir)

        def log_msg(self, msg):
            self.txt_log.insert(tk.END, msg + "\n")
            self.txt_log.see(tk.END)

        def start_generation(self):
            tpl = self.entry_tpl.get().strip()
            if tpl and os.path.exists(tpl):
                self.template_path = tpl

            if not self.template_path or not os.path.exists(self.template_path):
                messagebox.showwarning("Avertissement", "Veuillez sélectionner un fichier template Excel.")
                return
            if not self.listes_files:
                messagebox.showwarning("Avertissement", "Veuillez sélectionner au moins un fichier de liste d'élèves.")
                return

            self.btn_launch.config(state="disabled")
            self.txt_log.delete("1.0", tk.END)
            self.log_msg("Démarrage de la génération par lots...")

            def worker_task():
                try:
                    parent_d = self.entry_dest.get().strip() or os.path.abspath("fichiers_eleves")
                    separateur = self.var_sep.get()
                    c3_format = self.var_c3.get()

                    total_fichiers = 0
                    all_data = []
                    for l_file in self.listes_files:
                        eleves = parse_student_list(l_file)
                        if eleves:
                            all_data.append((l_file, eleves))

                    total_items = sum(len(e) for _, e in all_data)
                    current_item = 0

                    for l_file, eleves in all_data:
                        nom_classe = eleves[0]["classe"] if eleves[0]["classe"] else os.path.splitext(os.path.basename(l_file))[0]
                        target_dir = os.path.join(parent_d, nom_classe) if len(all_data) > 1 else parent_d
                        os.makedirs(target_dir, exist_ok=True)

                        for e in eleves:
                            nom = e["nom"]
                            prenom = e["prenom"]
                            classe = e["classe"] if e["classe"] else nom_classe
                            parts = [p for p in [classe, nom, prenom] if p]
                            
                            nom_f = ("_".join(parts).replace(" ", "_") if separateur == "_" else " ".join(parts)) + ".xlsx"
                            c3_val = f"{nom} {prenom}" if c3_format == "nom_prenom" else f"{prenom} {nom}"

                            destination = os.path.join(target_dir, nom_f)
                            
                            wb = openpyxl.load_workbook(self.template_path)
                            ws = wb.active
                            ws['C3'] = c3_val.strip()
                            wb.save(destination)

                            current_item += 1
                            total_fichiers += 1
                            m = f"✓ [{current_item}/{total_items}] [{nom_classe}] Généré : {nom_f}"
                            self.after(0, self.log_msg, m)

                    self.after(0, self.on_finish, total_fichiers, len(all_data), os.path.abspath(parent_d))
                except Exception as err:
                    self.after(0, lambda: messagebox.showerror("Erreur", str(err)))
                    self.after(0, lambda: self.btn_launch.config(state="normal"))

            threading.Thread(target=worker_task, daemon=True).start()

        def on_finish(self, total_files, total_classes, parent_dir):
            self.btn_launch.config(state="normal")
            self.log_msg(f"\n✓ Succès : {total_files} fichier(s) généré(s) pour {total_classes} classe(s) !")
            messagebox.showinfo("Succès !", f"🎉 {total_files} fichier(s) généré(s) pour {total_classes} classe(s) !\n\n📁 Dossier : {parent_dir}")

    app = TkApp()
    app.mainloop()


# =============================================================================
# MAIN : DÉTECTION DU MOTEUR DISPONIBLE
# =============================================================================
def main():
    try:
        import PySide6
        run_pyside6_app()
        return
    except ImportError:
        pass

    try:
        import tkinter
        run_tkinter_app()
        return
    except ImportError:
        pass

    print("===============================================================")
    print(" ⚠️ DÉPENDANCE GUI MANQUANTE")
    print("===============================================================")
    print("L'interface graphique nécessite 'PySide6' ou 'tkinter'.\n")
    
    choix = input("👉 Souhaitez-vous installer PySide6 automatiquement ? [O/n] : ").strip().lower()
    if choix in ['', 'o', 'oui', 'y', 'yes']:
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "PySide6"])
        except Exception:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "PySide6", "--break-system-packages"])
        run_pyside6_app()
    else:
        sys.exit(1)

if __name__ == "__main__":
    main()
